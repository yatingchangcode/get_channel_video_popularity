import ntplib
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib.parse import urlparse
import yaml
from apiclient.discovery import build 
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

TYPE_CHANNEL = 1
TYPE_VIDEO = 2

YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"


class popularity:
    def __init__(self):
        self.api_key = ""
        self.channelfile = "channel.xlsx"
        self.channelurls = []
        self.totalChVideoIDs =[[]]
        self.ntpPreciseTime = ""

    def loadconfig(self, confile):
        config = self.get_config(confile)
        self.api_key = config['apikey']
        self.channelfile = config['channelxlsx']

    def getPreciseTime(self):
        try:
            c = ntplib.NTPClient()
            response = c.request('pool.ntp.org')
            ts = response.tx_time
            today = datetime.fromtimestamp(response.tx_time)
            app_date = datetime(year=2020,month=3,day=20) #setup a dateti    me object
            self.ntpPreciseTime = today.strftime("%y/%m/%d, %H:%M:%S")
        except:
            today = date.today()
            self.ntpPreciseTime = today.strftime("%y/%m/%d, %H:%M:%S")

    def get_config(self, yamlConfigFile):
        with open(yamlConfigFile) as file:                                 
            config = yaml.safe_load(file)
            return config

    def read_channel_xlsx(self, channelfile):
        wb = load_workbook(channelfile)
        sheet = wb.active
        max_column = sheet.max_column
        for i in range(1, max_column + 1):
            if sheet.cell(row=1, column=i).value == "channelurl" :
                    url_idx = i
        print ('channelurl_idx =', url_idx)
        print('\n')                                                                                
        self.channelurls = self.parse_videoID(sheet, url_idx, TYPE_CHANNEL)

    def get_channelID_from_url(self, url):
        query = urlparse(url)
        if query.hostname in ('www.youtube.com', 'youtube.com'):
            if query.path[:9] == '/channel/':
                return query.path.split('/')[2]
        # fail?
        return None
    
    def get_videoID_from_url(self, url):
        query = urlparse(url)
        if query.hostname == 'youtu.be':
            return query.path[1:]
        if query.hostname in ('www.youtube.com', 'youtube.com'):
            if query.path == '/watch':
                p = parse_qs(query.query)
                return p['v'][0]
            if query.path[:7] == '/embed/':
                return query.path.split('/')[2]
            if query.path[:3] == '/v/':
                return query.path.split('/')[2]
        # fail?
        return None
    
    def parse_videoID(self, sheet, url_idx, urltype):
        IDs = []
        max_row = sheet.max_row
        for i in range(2, max_row + 1):
            url = sheet.cell(row=i, column=url_idx).value
            if url:
                if urltype == TYPE_CHANNEL:
                    ID = self.get_channelID_from_url(url)
                else:
                    ID = self.get_videoID_from_url(url)
                IDs.append(ID)
        return IDs

    def get_channel_all_videos(self, youtube, channel_id):
        # get Uploads playlist id
        res = youtube.channels().list(id=channel_id, 
                part='contentDetails').execute()
        playlist_id = res['items'][0]['contentDetails']['relatedPlaylists']['uploads']
    
        videoIDs = []
        videoTitles = []
        next_page_token = None
        #print(playlist_id)
        while 1:
            res = youtube.playlistItems().list(playlistId=playlist_id, 
                    part='snippet', 
                    maxResults=50,
                    pageToken=next_page_token).execute()
            #videos.append(res['items'])
            next_page_token = res.get('nextPageToken')
            for vid in res['items']: #handle 50 items
                videoIDs.append(vid['snippet']['resourceId']['videoId'])
                videoTitles.append(vid['snippet']['title'])
            next_page_token = res.get('nextPageToken')
    
            if next_page_token is None:
                break

        channelTitle = vid['snippet']['channelTitle']
        return channelTitle, videoIDs, videoTitles

    def setEachChannelWB(self, channelTitle, videoIDs, videoTitles):
        wb = Workbook()
        sheet = wb.active
        channel_file = "channel_" + channelTitle + ".xlsx"
        sheet.column_dimensions[get_column_letter(1)].width = 15.0
        sheet.column_dimensions[get_column_letter(2)].width = 30.0
    
        sheet.cell(row=1, column=1).value = "videoID"
        sheet.cell(row=1, column=2).value = "videoTitle"

        rowIdx = 2
        for i in range(len(videoIDs)):
            sheet.cell(row=rowIdx, column=1).value = videoIDs[i]
            sheet.cell(row=rowIdx, column=2).value = videoTitles[i]
            rowIdx += 1
        wb.save(filename = channel_file)

    def initVideoWB(self, sheet):
        sheet.column_dimensions[get_column_letter(1)].width = 10.0
        sheet.column_dimensions[get_column_letter(2)].width = 30.0
        sheet.column_dimensions[get_column_letter(3)].width = 20.0
        sheet.column_dimensions[get_column_letter(4)].width = 10.0
        sheet.column_dimensions[get_column_letter(5)].width = 10.0
        sheet.column_dimensions[get_column_letter(6)].width = 10.0
        sheet.column_dimensions[get_column_letter(7)].width = 10.0
        sheet.column_dimensions[get_column_letter(8)].width = 10.0
        sheet.column_dimensions[get_column_letter(9)].width = 10.0
        sheet.column_dimensions[get_column_letter(10)].width = 10.0
        sheet.column_dimensions[get_column_letter(11)].width = 30.0
        sheet.column_dimensions[get_column_letter(12)].width = 30.0
    
        sheet.cell(row=1, column=1).value = "DATE"
        sheet.cell(row=1, column=2).value = "TITLE"
        sheet.cell(row=1, column=3).value = "URL"
        sheet.cell(row=1, column=4).value = "viewCount"
        sheet.cell(row=1, column=5).value = "likeCount"
        sheet.cell(row=1, column=6).value = "dislikeCount"
        sheet.cell(row=1, column=7).value = "favoriteCount"
        sheet.cell(row=1, column=8).value = "commentCount"
        sheet.cell(row=1, column=9).value = "subscriberCount"
        sheet.cell(row=1, column=10).value = "tag"
        sheet.cell(row=1, column=11).value = "description"
        sheet.cell(row=1, column=12).value = "contentDetails"


    def setDataToExcel(self, sheet, dateidx, result, channel_rlt):
        sheet.cell(row=dateidx, column=1).value = self.ntpPreciseTime
        sheet.cell(row=dateidx, column=2).value = result["snippet"]["title"]
        sheet.cell(row=dateidx, column=3).value = "https://www.youtube.com/watch?v=" + result["id"]
        sheet.cell(row=dateidx, column=4).value = result["statistics"]["viewCount"]
        sheet.cell(row=dateidx, column=5).value = result["statistics"]["likeCount"]
        sheet.cell(row=dateidx, column=6).value = result["statistics"]["dislikeCount"]
        sheet.cell(row=dateidx, column=7).value = result["statistics"]["favoriteCount"]
        try:
            sheet.cell(row=dateidx, column=8).value = result["statistics"]["commentCount"]
        except:
            sheet.cell(row=dateidx, column=8).value = 0 

        sheet.cell(row=dateidx, column=9).value = channel_rlt["statistics"]["subscriberCount"]
        try:
            sheet.cell(row=dateidx, column=10).value = str(result["snippet"]["tags"])
        except:
             sheet.cell(row=dateidx, column=10).value = ""
        sheet.cell(row=dateidx, column=11).value = result['snippet']['description'] 
        sheet.cell(row=dateidx, column=12).value = str(result['contentDetails']) 

    def check_dateidx(self, sheet):
        check = 1
        row_idx = 2
        while (check):
            if not sheet.cell(row=row_idx, column=1).value: #which is null
                check = 0
                return row_idx
            row_idx += 1

    def multiple_video_details(self, youtube, sheet, urlIDs): 
        dateidx = self.check_dateidx(sheet)
        print("dateidx %s" % dateidx)
        column_idx = 2
        self.getPreciseTime()
        for ID in urlIDs:
            list_videos_byid = youtube.videos().list( 
                        id = ID, 
                        part = "id, snippet, contentDetails, statistics", 
    	).execute()
    
        # extracting the results from search response 
            results = list_videos_byid.get("items", []) 
            # empty list to store video details 
            videos = []
    
            for result in results:
                column = get_column_letter(column_idx)
                #sheet.column_dimensions[column].width = 40.0
                result_str = ""
                result_statistics = result["statistics"]
                for key in result_statistics:
                    result_str += key + ' : ' + result_statistics[key] + '\n'
    
                #search channel subscribercount
                channel = youtube.channels().list(
                        part="statistics",
                        id= result["snippet"]["channelId"]
                ).execute()
    
                channel_rltss = channel.get("items", [])
                for channel_rlt in channel_rltss:
                    result_str += 'subscriberCount' + ' : ' + channel_rlt["statistics"]["subscriberCount"]
                #======================
                self.setDataToExcel(sheet, dateidx, result, channel_rlt)
                dateidx += 1
                column_idx = column_idx + 1
    
    def setVideofEachChannelWB(self, youtube, channelTitle, videoIDs):
        pop_file = "video_" + channelTitle + ".xlsx"
        if os.path.exists(pop_file):
            wb = load_workbook(pop_file)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            self.initVideoWB(sheet)
        self.initVideoWB(sheet)
        self.multiple_video_details(youtube, sheet, videoIDs)
        wb.save(filename = pop_file)

def readpopularity():
    #config = get_config("apikey.conf")
    popu = popularity()
    popu.loadconfig("apikey.conf")
    popu.read_channel_xlsx(popu.channelfile)
    
    # creating youtube resource object for interacting with API 
    youtube = build(YOUTUBE_API_SERVICE_NAME, 
            YOUTUBE_API_VERSION, 
            developerKey = popu.api_key) 

    for channelid in popu.channelurls:
        channelTitle, videoIDs, videoTitles = popu.get_channel_all_videos(youtube, channelid)
        popu.setEachChannelWB(channelTitle, videoIDs, videoTitles)
        popu.setVideofEachChannelWB(youtube, channelTitle, videoIDs)

def main():
    readpopularity()

if __name__ == "__main__":
    main()

